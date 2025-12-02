import openpyxl
from utils.file_utils import disqualify_lead
from utils.email_utils import extract_root_domain
from rapidfuzz import fuzz

def apply_campaign_domain_suppression(
    wb, ws, suppression_file_path, domain_col_name, fuzzy_threshold=90, fuzzy_warning=75
):
    """
    Disqualify leads whose (root) domain matches or strongly fuzzy-matches
    any (root) domain in the campaign suppression list.
    If fuzzy match is between 75 and 89, log in a column for QA review.
    """
    sup_wb = openpyxl.load_workbook(suppression_file_path, read_only=True)
    sup_ws = sup_wb.active

    # Find suppression domain column index
    headers = [str(cell.value).strip().lower() for cell in sup_ws[1]]
    try:
        idx = headers.index(domain_col_name.strip().lower()) + 1
    except ValueError:
        raise ValueError(f"Suppression file is missing column '{domain_col_name}'")

    # Build set of suppression root domains and original mapping
    sup_root_domains = set()
    original_domain_map = {}
    for row in sup_ws.iter_rows(min_row=2):
        raw = row[idx - 1].value
        if raw:
            norm_root = extract_root_domain(str(raw).strip().lower())
            if norm_root:
                sup_root_domains.add(norm_root)
                original_domain_map[norm_root] = str(raw).strip()

    # Main sheet column indexes
    main_hdrs = [str(cell.value).strip().lower() for cell in ws[1]]
    col_status = main_hdrs.index('lead status') + 1
    col_reason = main_hdrs.index('dq reason') + 1
    col_comment = main_hdrs.index('qa comment') + 1
    col_domain = main_hdrs.index('domain') + 1

    # --- Add "Company Domain Fuzzy Suppression Score" column if not present ---
    qa_warning_header = "company domain fuzzy suppression score"
    if qa_warning_header in main_hdrs:
        qa_token_col = main_hdrs.index(qa_warning_header) + 1
    else:
        qa_token_col = ws.max_column + 1
        ws.cell(row=1, column=qa_token_col, value='Company Domain Fuzzy Suppression Score')

    # Iterate main rows
    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        dom_raw = str(row[col_domain - 1].value or '').strip().lower()
        dom_root = extract_root_domain(dom_raw)

        # 1. Exact/normalized root domain match (DQ)
        if dom_root in sup_root_domains:
            disqualify_lead(
                ws,
                row=r,
                col_status=col_status,
                col_reason=col_reason,
                col_comment=col_comment,
                reason='Client Suppression',
                comment=f"Domain '{dom_raw}' suppressed (Matched to Client Domain Exclusion '{original_domain_map[dom_root]}')"
            )
            continue

        # 2. Fuzzy root domain match
        warning_msgs = []
        dq_flag = False
        for sup_root in sup_root_domains:
            score = fuzz.ratio(dom_root, sup_root)
            if score >= fuzzy_threshold:
                disqualify_lead(
                    ws,
                    row=r,
                    col_status=col_status,
                    col_reason=col_reason,
                    col_comment=col_comment,
                    reason='Campaign Domain Suppression',
                    comment=f"Domain '{dom_raw}' suppressed (Fuzzy match to '{original_domain_map[sup_root]}', Score={score})"
                )
                dq_flag = True
                break  # Only need to DQ once
            elif fuzzy_warning <= score < fuzzy_threshold:
                warning_msgs.append(
                    f"Possible fuzzy match to suppression '{original_domain_map[sup_root]}' (Score={score})"
                )
        if dq_flag:
            continue

        # Write all warnings (semicolon separated) if any
        if warning_msgs:
            ws.cell(row=r, column=qa_token_col).value = "; ".join(warning_msgs)

    return wb, ws
