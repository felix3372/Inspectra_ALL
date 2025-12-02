import openpyxl
import re
from utils.file_utils import disqualify_lead
from rapidfuzz import fuzz
from collections import Counter

# --- Comprehensive stopwords ---
GENERIC_COMPANY_WORDS = {
    # Legal/company types
    "ltd", "limited", "pty", "inc", "corp", "corporation", "company", "plc", "group", "holdings",
    "sa", "nv", "llc", "gmbh", "co", "spa", "ag", "bv", "kg", "as", "ab", "srl", "sl", "oy", "sas",
    "sasu", "sarl", "ulc", "bhd", "a/s", "ap", "aps", "spzoo", "spolka", "zoo", "kk", "kabushiki",
    "kaisha", "aktiengesellschaft", "oyj", "incorporated", "sociedade", "anonima", "sae", "akciju",
    "akcine", "osakeyhtio", "yhtiö", "tov", "doo", "sro", "ae", "asa", "pte", "teo", "d.o.o.",
    # Geography (common)
    "new", "zealand", "australia", "international", "america", "american", "africa", "india", "china",
    "canada", "bahrain", "singapore", "europe", "france", "germany", "uk", "england", "scotland",
    "ireland", "russia", "russian", "emirates", "uae", "united", "states", "japan", "brazil",
    # Industry/material
    "airport", "resources", "media", "entertainment", "radio", "pharmaceuticals", "milk", "steel",
    "energy", "mining", "bank", "food", "chemicals", "insurance", "automotive", "transport",
    "logistics", "services", "systems", "technologies", "technology", "solutions", "infrastructure",
    "hospital", "aerospace", "investment", "management", "manufacturing", "electronics", "telecom",
    # Language
    "e", "de", "and", "the", "of", "for", "with", "a", "an", "et", "und", "en", "la", "le", "des",
    "das", "do", "da", "del", "di", "du", "su", "na", "to", "by", "on", "in", "at", "el",
    # Short/junk
    "co", "pen", "edge", "blue", "bsc", "sgps", "s.a.", "s.a", "pty", "pt", "inc.", "inc"
}

# --- Augment with world cities and countries if libraries installed ---
# try:
#     import geonamescache
#     gc = geonamescache.GeonamesCache()
#     cities = set(c['name'].lower() for c in gc.get_cities().values())
#     GENERIC_COMPANY_WORDS.update(cities)
# except ImportError:
#     pass

try:
    import pycountry
    countries = set(c.name.lower() for c in pycountry.countries)
    GENERIC_COMPANY_WORDS.update(countries)
except ImportError:
    pass

def normalize_company(name):
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

def get_all_tokens(name):
    if not name:
        return set()
    return set(
        w for w in re.split(r'[^a-z0-9]+', name.lower())
        if w and w not in GENERIC_COMPANY_WORDS
    )

def apply_campaign_company_suppression(
    wb, ws, suppression_file_path, company_col_name, fuzzy_high=90, fuzzy_low=75, fuzzy_warnings_limit=5
):
    """
    - Exact/normalized match: DQ
    - Fuzzy > 90: warn in 'Company name fuzzy suppression score (> 90)'
    - Fuzzy 75–89: warn (top 5) in 'Company name fuzzy suppression score'
    - Rare token matches: warn in 'Company name fuzzy suppression score' (with matched company names)
    """
    sup_wb = openpyxl.load_workbook(suppression_file_path, read_only=True)
    sup_ws = sup_wb.active

    headers = [str(cell.value).strip().lower() for cell in sup_ws[1]]
    try:
        idx = headers.index(company_col_name.strip().lower()) + 1
    except ValueError:
        raise ValueError(f"Suppression file is missing column '{company_col_name}'")

    sup_companies = set()
    normalized_to_original = dict()
    all_supp_tokens = set()
    company_tokens_map = dict()

    for row in sup_ws.iter_rows(min_row=2):
        raw = row[idx - 1].value
        norm = normalize_company(raw)
        tokens = get_all_tokens(str(raw))
        if norm:
            sup_companies.add(norm)
            normalized_to_original[norm] = str(raw).strip()
            all_supp_tokens.update(tokens)
            company_tokens_map[norm] = tokens

    # --- Count token frequencies across suppression list ---
    supp_token_counts = Counter()
    for tokens in company_tokens_map.values():
        supp_token_counts.update(tokens)
        
    COMMON_TOKEN_THRESHOLD = 5  # Only warn for tokens that appear <=5 times in suppression list

    main_hdrs = [str(cell.value).strip().lower() for cell in ws[1]]
    col_status = main_hdrs.index('lead status') + 1
    col_reason = main_hdrs.index('dq reason') + 1
    col_comment = main_hdrs.index('qa comment') + 1
    try:
        col_company = main_hdrs.index('company name') + 1
    except ValueError:
        col_company = main_hdrs.index('company') + 1

    # --- Add warning columns if not present ---
    warn_colname = "company name fuzzy suppression score"
    high_fuzzy_colname = "company name fuzzy suppression score (> 90)"
    if warn_colname in main_hdrs:
        qa_token_col = main_hdrs.index(warn_colname) + 1
    else:
        qa_token_col = ws.max_column + 1
        ws.cell(row=1, column=qa_token_col, value=warn_colname)
    # for > 90 fuzzy
    if high_fuzzy_colname in main_hdrs:
        qa_high_fuzzy_col = main_hdrs.index(high_fuzzy_colname) + 1
    else:
        qa_high_fuzzy_col = ws.max_column + 2 if qa_token_col == ws.max_column + 1 else ws.max_column + 1
        ws.cell(row=1, column=qa_high_fuzzy_col, value=high_fuzzy_colname)

    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        comp_raw = str(row[col_company - 1].value or '').strip()
        comp_norm = normalize_company(comp_raw)
        lead_tokens = get_all_tokens(comp_raw)

        warning_msgs = []
        high_fuzzy_msgs = []

        # 1. Exact/normalized match (DQ)
        if comp_norm in sup_companies:
            disqualify_lead(
                ws,
                row=r,
                col_status=col_status,
                col_reason=col_reason,
                col_comment=col_comment,
                reason='Client Suppression',
                comment=f"Company '{comp_raw}' suppressed (Exact/normalized match to Client Company Exclusion'{normalized_to_original[comp_norm]}')"
            )
            continue

        # 2. Fuzzy match: just warn, never DQ
        fuzzy_scores = []
        for sup_norm in sup_companies:
            score = fuzz.ratio(comp_norm, sup_norm)
            if score > fuzzy_high:
                high_fuzzy_msgs.append(
                    f"High fuzzy match to suppression '{normalized_to_original[sup_norm]}' (Score={score})"
                )
            elif fuzzy_low <= score <= fuzzy_high:
                fuzzy_scores.append((score, normalized_to_original[sup_norm]))
        # Only top N fuzzy warnings (75–90)
        if fuzzy_scores:
            fuzzy_scores.sort(reverse=True)
            for i, (score, name) in enumerate(fuzzy_scores[:fuzzy_warnings_limit]):
                warning_msgs.append(
                    f"Possible fuzzy match to suppression '{name}' (Score={score})"
                )
            if len(fuzzy_scores) > fuzzy_warnings_limit:
                warning_msgs.append(
                    f"...and {len(fuzzy_scores) - fuzzy_warnings_limit} more fuzzy matches."
                )

        # 3. Token (group/parent/brand) match (WARN ONLY, do not DQ) with company names
        rare_shared_tokens = {tok for tok in lead_tokens & all_supp_tokens if supp_token_counts[tok] <= COMMON_TOKEN_THRESHOLD}
        if rare_shared_tokens:
            for token in rare_shared_tokens:
                matched_companies = [
                    normalized_to_original[norm]
                    for norm, toks in company_tokens_map.items()
                    if token in toks
                ]
                warning_msgs.append(
                    f"Shared rare token '{token}' with suppression: {matched_companies}"
                )

        # Write all warnings if any
        if warning_msgs:
            ws.cell(row=r, column=qa_token_col).value = "; ".join(warning_msgs)
        if high_fuzzy_msgs:
            ws.cell(row=r, column=qa_high_fuzzy_col).value = "; ".join(high_fuzzy_msgs)

    return wb, ws
