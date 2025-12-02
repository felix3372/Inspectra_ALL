from openpyxl.styles import PatternFill

def convert_sales_link(wb, ws):
    """
    Converts LinkedIn Sales Navigator URLs in the 'Linkedin Link Prospect' column
    to standard public profile URLs.
    Also highlights cells containing 'ACW', 'ACw', or 'acw'.
    """
    # Define a highlight fill (yellow)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Normalize headers to find the column
    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    try:
        linkedin_col = headers.index("linkedin link prospect") + 1
    except ValueError:
        raise ValueError("❌ Column 'Linkedin Link Prospect' not found.")

    for row in ws.iter_rows(min_row=2):
        cell = row[linkedin_col - 1]
        if not cell.value:
            continue

        original = str(cell.value).strip()

        # Clean the Sales Navigator URL
        if "/sales/people/" in original:
            trimmed = original.split(",")[0]
            cleaned = trimmed.replace("/sales/people/", "/in/")
            cell.value = cleaned

        # Highlight if 'acw' is in the cleaned link (case-insensitive)
        if "acw" in str(cell.value).lower():
            cell.fill = highlight_fill

    return wb, ws
