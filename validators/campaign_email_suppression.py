import openpyxl
from utils.file_utils import disqualify_lead
from utils.email_utils import generate_email_permutations, extract_root_domain

def normalize_email_root(email):
    """
    Lowercase, strip whitespace, and replace the domain with its root.
    e.g., 'ahmad@maaden.com.sa' -> 'ahmad@maaden'
    """
    if not email or '@' not in email:
        return (email or "").lower().strip()
    local, domain = email.lower().strip().split('@', 1)
    root = extract_root_domain(domain)
    return f"{local}@{root}"

def apply_campaign_email_suppression(wb, ws, suppression_file_path, email_col_name):
    """
    Disqualify leads by checking:
    1. FIRST: Check if the actual email in the lead matches suppression list
    2. THEN: Check if any generated email permutations match suppression list
    
    Both checks use root domain logic for matching.
    """
    # Load suppression workbook and worksheet
    sup_wb = openpyxl.load_workbook(suppression_file_path, read_only=True)
    sup_ws = sup_wb.active

    # Identify suppression column index
    headers = [str(cell.value).strip() for cell in sup_ws[1]]
    try:
        sup_email_idx = headers.index(email_col_name) + 1
    except ValueError:
        raise ValueError(f"Suppression file is missing column '{email_col_name}'")

    # Build set of normalized suppression emails (root-domain) AND mapping to originals
    sup_emails = set()
    suppression_email_map = {}  # normalized -> original email
    
    for row in sup_ws.iter_rows(min_row=2):
        raw = row[sup_email_idx - 1].value
        if raw:
            original_email = str(raw).strip()
            norm = normalize_email_root(original_email)
            sup_emails.add(norm)
            suppression_email_map[norm] = original_email  # Store mapping

    # Main sheet column indexes
    main_hdrs = [str(cell.value).strip().lower() for cell in ws[1]]
    col_status = main_hdrs.index('lead status') + 1
    col_reason = main_hdrs.index('dq reason') + 1
    col_comment = main_hdrs.index('qa comment') + 1
    col_first = main_hdrs.index('first name') + 1
    col_last = main_hdrs.index('last name') + 1
    col_domain = main_hdrs.index('domain') + 1
    
    # Try to find email address column (optional - may not exist in all sheets)
    col_email = None
    try:
        col_email = main_hdrs.index('email address') + 1
    except ValueError:
        # If 'email address' column doesn't exist, try 'email'
        try:
            col_email = main_hdrs.index('email') + 1
        except ValueError:
            pass  # No email column found, will only check permutations

    # Iterate main rows
    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        first = str(row[col_first - 1].value or '').strip()
        last = str(row[col_last - 1].value or '').strip()
        domain = str(row[col_domain - 1].value or '').strip()
        
        # Get actual email if column exists
        actual_email = None
        if col_email:
            actual_email = str(row[col_email - 1].value or '').strip()
        
        matched = False
        matched_email = None
        
        # ===== STEP 1: Check actual email FIRST (if it exists) =====
        if actual_email and '@' in actual_email:
            norm_actual = normalize_email_root(actual_email)
            if norm_actual in sup_emails:
                matched = True
                matched_email = suppression_email_map.get(norm_actual, norm_actual)
        
        # ===== STEP 2: If actual email didn't match, check permutations =====
        if not matched:
            # Generate permutations from first name, last name, and domain
            perms = generate_email_permutations(first, last, domain)
            
            # Check each permutation using root-domain logic
            for perm in perms:
                norm_perm = normalize_email_root(perm)
                if norm_perm in sup_emails:
                    matched = True
                    matched_email = suppression_email_map.get(norm_perm, norm_perm)
                    break  # Only need to flag once
        
        # ===== STEP 3: Disqualify if any match was found =====
        if matched:
            disqualify_lead(
                ws,
                row=r,
                col_status=col_status,
                col_reason=col_reason,
                col_comment=col_comment,
                reason='Client Suppression',
                comment=f'Email Found in Exclusion List (Root Domain Match "{matched_email}")',
            )

    return wb, ws