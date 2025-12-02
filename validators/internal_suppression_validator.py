import openpyxl
from utils.file_utils import disqualify_lead
from utils.email_utils import generate_email_permutations

def apply_internal_suppression(wb, ws, suppression_file_path, suppression_info):
    """
    Disqualifies leads from the main worksheet if they appear in the suppression list.
    Matches by:
      1. FIRST: Actual email address in the lead (if exists)
      2. THEN: Any known email pattern based on First + Last + Domain (permutations)
      3. Composite key: first 3 chars of First + Last + Domain (lowercased)
    """

    # Load suppression file
    suppression_wb = openpyxl.load_workbook(suppression_file_path)
    sheet_name = suppression_info.get("sheet_name")
    if not sheet_name or sheet_name not in suppression_wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in suppression file.")
    suppression_ws = suppression_wb[sheet_name]

    # Extract column names from mappings
    mappings = suppression_info["map"]
    file_name = suppression_info["file_name"]

    suppression_headers = [str(cell.value).strip().lower() for cell in suppression_ws[1]]
    suppression_col_indexes = {
        key: suppression_headers.index(mappings[key].strip().lower()) + 1
        for key in ["first_name", "last_name", "email", "domain"]
    }

    # Get main file column indexes
    headers = [str(cell.value).strip().lower() for cell in ws[1]]
    col_status = headers.index("lead status") + 1
    col_reason = headers.index("dq reason") + 1
    col_comment = headers.index("qa comment") + 1
    col_email = headers.index("email address") + 1
    col_first = headers.index("first name") + 1
    col_last = headers.index("last name") + 1
    col_domain = headers.index("domain") + 1

    # Build suppression email patterns and composite set
    suppression_email_set = set()
    composite_set = set()

    for row in suppression_ws.iter_rows(min_row=2):
        first = str(row[suppression_col_indexes["first_name"] - 1].value or "")
        last = str(row[suppression_col_indexes["last_name"] - 1].value or "")
        domain = str(row[suppression_col_indexes["domain"] - 1].value or "")
        email = str(row[suppression_col_indexes["email"] - 1].value or "").strip().lower()

        # Generate email patterns from suppression entries
        suppression_email_set.update(generate_email_permutations(first, last, domain))

        # Also include literal email if provided
        if email:
            suppression_email_set.add(email)

        # Add composite key
        fn = first[:3].lower()
        ln = last[:3].lower()
        dm = domain[:3].lower()
        if fn and ln and dm:
            composite_set.add(fn + ln + dm)

    # Apply suppression to main data
    for row in ws.iter_rows(min_row=2):
        row_num = row[0].row
        actual_email = str(row[col_email - 1].value or "").strip().lower()
        first = str(row[col_first - 1].value or "")
        last = str(row[col_last - 1].value or "")
        domain = str(row[col_domain - 1].value or "")

        # Generate composite key for row
        fn = first[:3].lower()
        ln = last[:3].lower()
        dm = domain[:3].lower()
        composite = fn + ln + dm

        matched = False
        
        # ===== STEP 1: Check actual email FIRST =====
        if actual_email and actual_email in suppression_email_set:
            matched = True
        
        # ===== STEP 2: If actual email didn't match, check permutations =====
        if not matched:
            # Generate permutations from lead's first, last, domain
            lead_permutations = generate_email_permutations(first, last, domain)
            
            # Check if any permutation matches suppression set
            if any(perm.lower() in suppression_email_set for perm in lead_permutations):
                matched = True
        
        # ===== STEP 3: Check composite key =====
        if not matched and composite in composite_set:
            matched = True
        
        # ===== STEP 4: Disqualify if any match was found =====
        if matched:
            disqualify_lead(
                ws,
                row=row_num,
                col_status=col_status,
                col_reason=col_reason,
                col_comment=col_comment,
                reason="Internal Suppression",
                comment=f'Lead found in Internal "{file_name}"'
            )

    return wb, ws