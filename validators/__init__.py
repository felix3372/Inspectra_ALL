import openpyxl
import tempfile
import streamlit as st
import time
from datetime import datetime

# Core validators
# REMOVED: from .header_validator import validate_required_headers  # Now handled in Home.py
# REMOVED: from .country_name_validator import find_first_invalid_country  # Now handled in Home.py
from .cleaner_validator import clean_data
from .email_validator import validate_emails
from .first_last_name_validator import first_last_name_validator
from .convert_sales_link import convert_sales_link
from .normalize_revenue_size_validator import normalize_revenue_size_validator
from .same_client_validator import same_client_validator
from .internal_suppression_validator import apply_internal_suppression
from .normalize_linkedin_link_validator import normalize_linkedin_link_validator  # normalize linkedin link
from .normalize_linkedin_company_link import normalize_linkedin_company_link  # normalize company linkedin link
from .validate_prospect_link_name_match import validate_prospect_link_name_match  # check prospect's name in link
from .validate_company_domain_match import validate_company_domain_match  # check is company has a valid domain
from .domain_country_validator import domain_country_validator  # Check Domain Geo
from .job_level_validator import job_level_validator  # add job level to job title
from .country_code_prefix_validator import country_code_validator
from .tollfree_validator import tollfree_validator
from .validate_phone_by_country import validate_phone_by_country  # Validate phone numbers by country

# Separate campaign suppression validators
from .campaign_email_suppression import apply_campaign_email_suppression
from .campaign_company_suppression import apply_campaign_company_suppression
from .campaign_domain_suppression import apply_campaign_domain_suppression


def format_time_elapsed(seconds):
    """
    Format elapsed time in a human-readable way.
    
    Args:
        seconds: Number of seconds elapsed
    
    Returns:
        Formatted string like "2m 15s" or "1h 5m 30s"
    """
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes}m {secs}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours}h {minutes}m {secs}s"


def estimate_remaining_time(elapsed, current_step, total_steps):
    """
    Estimate remaining time based on current progress.
    
    Args:
        elapsed: Seconds elapsed so far
        current_step: Current step number
        total_steps: Total number of steps
    
    Returns:
        Estimated seconds remaining
    """
    if current_step == 0:
        return 0
    
    avg_time_per_step = elapsed / current_step
    remaining_steps = total_steps - current_step
    return avg_time_per_step * remaining_steps


def run_all_validations(uploaded_file,
                        suppressed_input,
                        suppression_info=None,
                        campaign_suppression_info=None):
    """
    Runs the full validation pipeline:
      - Core data checks & cleaning
      - Internal suppression (if provided)
      - Campaign-wise suppression by email, company, and/or domain (if provided)
      - Other field & format validators
    
    NOTE: Country validation and header validation are now performed on file upload in Home.py
    
    Args:
      uploaded_file: file-like object or path to main Excel workbook
      suppressed_input: string of end-client names
      suppression_info: dict for internal suppression (or None)
      campaign_suppression_info: tuple (supp_file, modes_dict) or None
    
    Returns:
      Path to the validated Excel file
    """
    start_time = time.time()
    
    # Load the main workbook
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    # Get total rows for display
    total_rows = ws.max_row - 1  # Subtract header row
    
    # Extract campaign suppression args
    camp_file = None
    camp_modes = {}
    if campaign_suppression_info:
        camp_file, camp_modes = campaign_suppression_info
    
    # Create UI elements for progress tracking
    progress_bar = st.progress(0)
    
    # Create columns for better layout
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        progress_text = st.empty()
    with col2:
        time_text = st.empty()
    with col3:
        eta_text = st.empty()
    
    # Single status line that updates
    status_line = st.empty()
    
    # Build list of validators to run
    # NOTE: Country and Header validators removed - now run on file upload in Home.py
    validators = [
        (clean_data, True, [], "Cleaning Data", 3),
        (same_client_validator, True, [suppressed_input], "Checking Client Names", 2),
        (apply_internal_suppression, True, [suppression_info["path"], suppression_info], 
         "Applying Internal Suppression", 5) if suppression_info else None,
    ]
    
    # Campaign suppression modes with estimated weights
    if camp_file and camp_modes:
        if camp_modes.get("email"):
            validators.append(
                (apply_campaign_email_suppression, True,
                 [camp_file, camp_modes.get("email")], 
                 "Applying Email Suppression", 4)
            )
        if camp_modes.get("company"):
            validators.append(
                (apply_campaign_company_suppression, True,
                 [camp_file, camp_modes.get("company")],
                 "Applying Company Suppression", 4)
            )
        if camp_modes.get("domain"):
            validators.append(
                (apply_campaign_domain_suppression, True,
                 [camp_file, camp_modes.get("domain")],
                 "Applying Domain Suppression", 4)
            )
    
    # Core validations with time weights (relative processing time)
    validators += [
        (validate_emails, True, [], "Validating Email Formats", 5),
        (domain_country_validator, True, [], "Validating Domain Geography", 3),
        (first_last_name_validator, True, [], "Validating Names", 2),
        (convert_sales_link, True, [], "Converting Sales Links", 2),
        (normalize_linkedin_link_validator, True, [], "Normalizing LinkedIn Links", 2),
        (normalize_linkedin_company_link, True, [], "Normalizing Company Links", 2),
        (validate_company_domain_match, True, [], "Validating Company Domains", 3),
        (validate_prospect_link_name_match, True, [], "Validating Prospect Links", 3),
        (normalize_revenue_size_validator, True, [], "Normalizing Revenue Data", 2),
        (job_level_validator, True, [], "Processing Job Levels", 2),
        (country_code_validator, True, [], "Validating Country Codes", 2),
        (tollfree_validator, True, [], "Checking Toll-Free Numbers", 1),
        (validate_phone_by_country, True, [], "Validating Phone Numbers", 3),
    ]
    
    # Remove None entries
    validators = [v for v in validators if v]
    total = len(validators)
    
    # Calculate total weight for accurate progress
    total_weight = sum(v[4] for v in validators)
    completed_weight = 0
    
    # Track validation metrics
    validation_metrics = {
        "start_time": datetime.now(),
        "validators_run": [],
        "rows_processed": total_rows
    }
    
    # Execute validators in order
    for idx, validator_tuple in enumerate(validators, start=1):
        validator, expects_ws, args, display_name, weight = validator_tuple
        
        # Start timing this validator
        validator_start = time.time()
        
        # Update progress display
        progress_text.text(f"🔄 {display_name}... ({idx}/{total})")
        
        # Update time elapsed
        elapsed = time.time() - start_time
        time_text.text(f"⏱️ {format_time_elapsed(elapsed)}")
        
        # Estimate remaining time
        if idx > 1:  # Only estimate after first validator
            eta_seconds = estimate_remaining_time(elapsed, completed_weight, total_weight)
            eta_text.text(f"⏳ ETA: {format_time_elapsed(eta_seconds)}")
        
        # Run validator
        try:
            if expects_ws:
                wb, ws = validator(wb, ws, *args)
            else:
                validator(ws, *args)
            
            # Track successful validation
            validator_time = time.time() - validator_start
            validation_metrics["validators_run"].append({
                "name": display_name,
                "time": validator_time
            })
            
        except Exception as e:
            # Handle errors gracefully
            progress_text.empty()
            progress_bar.empty()
            time_text.empty()
            eta_text.empty()
            st.error(f"❌ Error in {display_name}: {str(e)}")
            return None
        
        completed_weight += weight
        progress_bar.progress(min(completed_weight / total_weight, 1.0))
    
    # Clear progress indicators
    progress_text.empty()
    time_text.empty()
    eta_text.empty()
    progress_bar.empty()
    
    # Calculate final metrics
    total_elapsed = time.time() - start_time
    validation_metrics["end_time"] = datetime.now()
    validation_metrics["total_time"] = total_elapsed
    
    # Analyze the results for basic counts
    headers = [cell.value for cell in ws[1]]
    lead_status_col = headers.index("Lead Status") + 1 if "Lead Status" in headers else None
    
    # Count disqualified and valid leads
    disqualified = 0
    valid_leads = 0
    
    if lead_status_col:
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row, lead_status_col).value
            if status and "Disqualified" in str(status):
                disqualified += 1
            else:
                valid_leads += 1
    else:
        # If no Lead Status column, assume all are valid
        valid_leads = total_rows
    
    # Display completion summary
    st.success("🎉 All validations completed successfully!")
    
    # Show analysis summary
    st.markdown("### 📊 Data Analysis Summary")
    
    # Main metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Total Rows", 
            f"{total_rows:,}",
            delta=None
        )
    
    with col2:
        st.metric(
            "Valid Leads", 
            f"{valid_leads:,}",
            delta=f"{(valid_leads/total_rows*100):.1f}%" if total_rows > 0 else "0%",
            delta_color="normal"
        )
    
    with col3:
        st.metric(
            "Disqualified", 
            f"{disqualified:,}",
            delta=f"{(disqualified/total_rows*100):.1f}%" if total_rows > 0 else "0%",
            delta_color="inverse"
        )
    
    with col4:
        pass_rate = (valid_leads/total_rows*100) if total_rows > 0 else 0
        st.metric(
            "Pass Rate", 
            f"{pass_rate:.1f}%",
            delta="Good" if pass_rate > 80 else "Review needed" if pass_rate > 60 else "Poor",
            delta_color="normal" if pass_rate > 60 else "inverse"
        )
    
    # Processing details in expander
    with st.expander("⚙️ Processing Details", expanded=False):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "Total Time", 
                format_time_elapsed(total_elapsed),
                delta=None
            )
        
        with col2:
            st.metric(
                "Validators Run", 
                f"{len(validators)}",
                delta=None
            )
        
        with col3:
            avg_time_per_row = (total_elapsed / total_rows) if total_rows > 0 else 0
            st.metric(
                "Avg Time/Row", 
                f"{avg_time_per_row:.3f}s",
                delta=None
            )
        
        with col4:
            rows_per_minute = (total_rows / total_elapsed * 60) if total_elapsed > 0 else 0
            st.metric(
                "Processing Speed", 
                f"{rows_per_minute:.0f} rows/min",
                delta=None
            )
        
        # Show time breakdown for longer runs
        if total_elapsed > 10 and len(validation_metrics["validators_run"]) > 0:
            st.divider()
            st.subheader("Time Breakdown (Top 5)")
            
            # Sort validators by time taken
            sorted_validators = sorted(
                validation_metrics["validators_run"], 
                key=lambda x: x["time"], 
                reverse=True
            )
            
            # Display as a simple table
            for v in sorted_validators[:5]:  # Show top 5 time consumers
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.text(v["name"])
                with col2:
                    st.text(format_time_elapsed(v["time"]))
    
    # Save result
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()
    
    # Final status message with quality indicator
    if pass_rate >= 80:
        quality_emoji = "🌟"
        quality_text = "Excellent data quality!"
    elif pass_rate >= 60:
        quality_emoji = "👍"
        quality_text = "Good data quality."
    else:
        quality_emoji = "⚠️"
        quality_text = "Data needs review."
    
    st.info(f"{quality_emoji} {quality_text} Processed {total_rows:,} rows with {pass_rate:.1f}% pass rate in {format_time_elapsed(total_elapsed)}")
    
    return temp_file.name