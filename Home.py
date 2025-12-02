import io
import os
import glob
import tempfile
import openpyxl
import streamlit as st

from validators import run_all_validations
from validators.country_name_validator import find_first_invalid_country
from validators.header_validator import validate_required_headers  # Import header validator
from utils.ui_helpers import show_splash
from utils.file_utils import build_pycountry_csv
from utils.config import SUPPRESSION_ROOT
from validators.duplicate_header_validator import validate_duplicate_headers


ALLOWED_EXTS = (".xlsx", ".xlsm")

st.set_page_config(page_title="Inspectra", layout="centered")

def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

local_css("styles/inspectra.css")

# ---- SHOW SPLASH ----
show_splash("animations/blueloader.json")

# --- HERO SECTION ---
st.markdown("""
<div class="inspectra-hero">
    <div class="inspectra-inline">
        <span class="inspectra-title">Inspectra</span>
        <span class="inspectra-divider">|</span>
        <span class="inspectra-tagline">Where Precision Meets Integrity.</span>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<p style='text-align: center; font-size: 1.09rem; font-weight: 500; color: #333; margin-top: -0.5rem; margin-bottom: 1.2rem;'>
    Better, Faster, and More Accurate QA Checks — done with just a click.
</p><br><br>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload your campaign or ABM file to get started", type=["xlsx"])

# Small helpers local to this file
def drive_available(path: str) -> bool:
    try:
        return os.path.isdir(path)
    except Exception:
        return False

@st.cache_resource(show_spinner=False)
def list_suppression_files(root: str):
    """Return list of (label, fullpath) for Excel files under root, newest first."""
    if not drive_available(root):
        return []
    files = glob.glob(os.path.join(root, "**", "*"), recursive=True)
    paths = [p for p in files if p.lower().endswith(ALLOWED_EXTS) and os.path.isfile(p)]
    paths.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    def label(p):
        parent = os.path.basename(os.path.dirname(p))
        return f"{parent}/{os.path.basename(p)}" if parent else os.path.basename(p)
    return [(label(p), p) for p in paths[:500]]

# --- EARLY VALIDATION CHECKS (Headers & Country) ---
file_is_valid = False
uploaded_bytes = None

if uploaded_file:
    # Cache the bytes once so we can safely reload workbooks multiple times
    uploaded_bytes = uploaded_file.getvalue()
    with st.spinner("Checking file format, headers, and data quality..."):
        try:
            wb_check = openpyxl.load_workbook(io.BytesIO(uploaded_bytes))
            ws_check = wb_check.active

            # Step 1: Check Required Headers FIRST
            try:
                validate_required_headers(ws_check)
                headers_valid = True
            except ValueError as e:
                # Headers are missing - show error and stop
                st.error(str(e))
                st.info(
                    "💡 **Tip:** Make sure your file includes all required columns. "
                    "You can add missing columns with blank values if needed."
                )
                st.stop()

            # Step 1.5: Check for Duplicate Headers ✅ ADD THIS NEW SECTION
            try:
                validate_duplicate_headers(ws_check)
            except ValueError as e:
                # Duplicate headers found - show error and stop
                st.error(str(e))
                st.stop()

            # Step 2: Country validation (only runs if headers pass)
            invalid_country, row_num = find_first_invalid_country(ws_check)

            if invalid_country and invalid_country != "Missing 'Country' column.":
                st.error(
                    f"❌ Whoops! There's a country name I don't recognize at row {row_num}: '{invalid_country}'.\n\n"
                    "Please use the full, official country name exactly as it appears in international standards.\n\n"
                    "✅ **Examples of accepted country names:**\n"
                    "• United States\n"
                    "• United Kingdom\n"
                    "• India\n"
                    "• Germany\n"
                    "• Australia\n\n"
                    "For the complete list, see: https://www.iban.com/country-codes\n\n"
                    "Once you've fixed the country name, please upload your file again. 👍"
                )
                with st.expander("Optional: Download the list of acceptable ISO country names"):
                    st.caption("Generated via pycountry (ISO 3166).")
                    st.download_button(
                        "Download ISO country names (CSV)",
                        data=build_pycountry_csv(),
                        file_name="acceptable_country_names.csv",
                        mime="text/csv"
                    )
                st.stop()
            elif invalid_country == "Missing 'Country' column.":
                st.warning("⚠️ Missing 'Country' column in the file. This column is recommended for proper validation.")
                file_is_valid = True
            else:
                # Both headers and countries are valid
                file_is_valid = True
                st.success("✅ File format check passed! All required headers present and country names are valid.")
                total_rows = ws_check.max_row - 1
                total_cols = ws_check.max_column
                st.info(f"📊 File loaded: {total_rows:,} rows, {total_cols} columns")

        except Exception as e:
            st.error(f"❌ Could not read the uploaded file: {e}")
            st.stop()

# Offer the ISO download even when things pass
if uploaded_file and file_is_valid:
    with st.expander("Need the list of acceptable country names? (Optional)"):
        st.caption("Download a CSV generated from pycountry (ISO 3166).")
        st.download_button(
            "Download ISO country names (CSV)",
            data=build_pycountry_csv(),
            file_name="acceptable_country_names.csv",
            mime="text/csv"
        )

# --- SUPPRESSION OPTIONS CARD ---
if uploaded_file and file_is_valid:
    with st.container():
        st.markdown("""
        <div class="section">
            <div class="section-title">🛡️ Suppression Options</div>
            <span style="color:#444;">Optionally, select and map a suppression file to enhance internal or campaign QA checks.</span>
        </div>
        """, unsafe_allow_html=True)
        col1, col2 = st.columns([0.05, 0.95])
        with col2:
            # --- Internal Suppression ---
            if "apply_internal" not in st.session_state:
                st.session_state.apply_internal = False
            st.session_state.apply_internal = st.checkbox(
                "Apply Internal Suppression", value=st.session_state.apply_internal
            )

            suppression_info = {}
            suppression_temp_path = None  # if we create a temp file for uploaded suppression

            if st.session_state.apply_internal:
                st.markdown("#### Internal Suppression File")

                # Let user choose source: network drive or upload (fallback to upload if drive missing)
                drive_ok = drive_available(SUPPRESSION_ROOT)
                if drive_ok:
                    source_mode = st.radio(
                        "Choose suppression source",
                        ["Network Drive", "Upload File"],
                        index=0,
                        help="If your Y: drive is mounted, pick a file from the dropdown. Otherwise upload a file."
                    )
                else:
                    st.warning("⚠️ Suppression drive not found. Please upload the suppression file instead.")
                    source_mode = "Upload File"

                suppression_file_bytes = None
                suppression_filename = None

                if source_mode == "Network Drive" and drive_ok:
                    options = list_suppression_files(SUPPRESSION_ROOT)
                    if not options:
                        st.info("No Excel files found under the suppression folder. Switch to **Upload File**.")
                    else:
                        labels = [lbl for (lbl, _) in options]
                        choice = st.selectbox("Select Internal Suppression File", labels)
                        if choice:
                            path = dict(options)[choice]
                            suppression_filename = os.path.basename(path)
                            # Read into memory for sheet header preview
                            try:
                                with open(path, "rb") as f:
                                    suppression_file_bytes = f.read()
                            except Exception as e:
                                st.error(f"❌ Could not read suppression file: {e}")
                                suppression_file_bytes = None

                            # We'll still pass a path to validators (likely expect a path)
                            suppression_path_for_validator = path
                else:
                    uploaded_supp = st.file_uploader(
                        "Upload suppression Excel (.xlsx/.xlsm)",
                        type=["xlsx", "xlsm"],
                        key="internal_supp_upload"
                    )
                    if uploaded_supp:
                        suppression_filename = uploaded_supp.name
                        suppression_file_bytes = uploaded_supp.getvalue()
                        # Create a temp file so existing validator code that expects a path will work
                        try:
                            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(suppression_filename)[1])
                            tmp.write(suppression_file_bytes)
                            tmp.flush(); tmp.close()
                            suppression_path_for_validator = tmp.name
                            suppression_temp_path = tmp.name
                        except Exception as e:
                            st.error(f"❌ Could not create temporary file for suppression: {e}")
                            suppression_file_bytes = None
                            suppression_path_for_validator = None

                # Build mapping UI if we have suppression bytes (for header preview)
                if suppression_file_bytes:
                    try:
                        suppression_wb = openpyxl.load_workbook(io.BytesIO(suppression_file_bytes), read_only=True)
                        sheet_names = suppression_wb.sheetnames
                        selected_sheet = st.selectbox("Select Sheet in Suppression File", sheet_names)
                        suppression_ws = suppression_wb[selected_sheet]

                        suppression_headers = ["Select an option"] + [
                            str(cell.value).strip() for cell in suppression_ws[1]
                        ]

                        st.markdown("#### Match Columns from Suppression File")
                        first_name_map = st.selectbox("Match 'First Name'", suppression_headers, key="map_fn")
                        last_name_map = st.selectbox("Match 'Last Name'", suppression_headers, key="map_ln")
                        email_map = st.selectbox("Match 'Email Address'", suppression_headers, key="map_email")
                        domain_map = st.selectbox("Match 'Domain'", suppression_headers, key="map_domain")

                        suppression_info = {
                            "path": suppression_path_for_validator,  # path for validator
                            "sheet_name": selected_sheet,
                            "map": {
                                "first_name": first_name_map,
                                "last_name": last_name_map,
                                "email": email_map,
                                "domain": domain_map
                            },
                            "file_name": suppression_filename
                        }

                    except Exception as e:
                        st.error(f"❌ Failed to load suppression file: {e}")

            # --- Divider for clarity ---
            st.markdown('<hr style="margin:0em 0 0.3em 0;border:1px solid #e5e5e5;">', unsafe_allow_html=True)

            # --- Campaign Wise Suppression (unchanged logic, still uses active sheet) ---
            apply_campaign_supp = st.checkbox("Apply Campaign Wise Suppression", key="campaign_supp_checkbox")
            campaign_suppression_modes = {}
            campaign_suppression_headers = []
            campaign_suppression_mapping = {}
            campaign_supp_file = None

            if apply_campaign_supp:
                campaign_supp_file = st.file_uploader(
                    "Upload Campaign Suppression File",
                    type=["xlsx"],
                    key="campaign_supp_file"
                )
                if campaign_supp_file:
                    campaign_supp_wb = openpyxl.load_workbook(campaign_supp_file, read_only=True)
                    campaign_supp_ws = campaign_supp_wb.active
                    campaign_suppression_headers = [str(cell.value).strip() for cell in campaign_supp_ws[1]]

                    st.markdown("#### Choose Suppression Types and Map Columns")
                    # Email Wise
                    apply_email_supp = st.checkbox("Apply Email Wise Suppression")
                    if apply_email_supp:
                        selected_email_col = st.selectbox(
                            "Map Email Column from Suppression File",
                            campaign_suppression_headers,
                            key="campaign_email_col"
                        )
                        if selected_email_col:
                            campaign_suppression_mapping["email"] = selected_email_col

                    # Company Name Wise
                    apply_company_supp = st.checkbox("Apply Company Name Wise Suppression")
                    if apply_company_supp:
                        selected_company_col = st.selectbox(
                            "Map Company Name Column from Suppression File",
                            campaign_suppression_headers,
                            key="campaign_company_col"
                        )
                        if selected_company_col:
                            campaign_suppression_mapping["company"] = selected_company_col

                    # Domain Wise
                    apply_domain_supp = st.checkbox("Apply Domain Wise Suppression")
                    if apply_domain_supp:
                        selected_domain_col = st.selectbox(
                            "Map Domain Column from Suppression File",
                            campaign_suppression_headers,
                            key="campaign_domain_col"
                        )
                        if selected_domain_col:
                            campaign_suppression_mapping["domain"] = selected_domain_col

                    # Consolidate selected modes
                    campaign_suppression_modes = {key: col for key, col in campaign_suppression_mapping.items() if col}
                    if not campaign_suppression_modes:
                        st.info("ℹ️ Please select at least one suppression type and map its column.")

# --- VALIDATION & RESULTS CARD ---
if uploaded_file and file_is_valid:
    with st.form("suppression_form"):
        suppressed_input = st.text_input(
            "Enter the End Client Name(s) to Proceed, seperate multiple names by comma (,).",
            placeholder="Amazon, AWS, Microsoft"
        )
        submitted = st.form_submit_button("⚙️ Run Full Validation")

    if submitted:
        campaign_supp_selected = 'apply_campaign_supp' in locals() and apply_campaign_supp
        campaign_supp_invalid = campaign_supp_selected and not campaign_suppression_modes

        # Additional guard: ensure internal mapping isn't left at placeholder
        def has_unmapped_placeholders(mapping: dict) -> bool:
            if not mapping:
                return True
            return any(v in (None, "", "Select an option") for v in mapping.values())

        if not suppressed_input.strip():
            st.warning("⚠️ Please enter at least one End Client name to proceed.")
        elif st.session_state.get("apply_internal") and (not suppression_info.get("map") or has_unmapped_placeholders(suppression_info.get("map"))):
            st.warning("⚠️ Please complete all internal suppression selections before proceeding.")
        elif campaign_supp_invalid:
            st.warning("⚠️ Please select at least one campaign suppression type and map its column before proceeding.")
        else:
            with st.spinner("Running all validations..."):
                suppression_temp_path = locals().get("suppression_temp_path", None)
                try:
                    # Pass a fresh BytesIO each time
                    validated_file_path = run_all_validations(
                        io.BytesIO(uploaded_bytes),
                        suppressed_input,
                        suppression_info if st.session_state.get("apply_internal") else None,
                        (campaign_supp_file, campaign_suppression_modes) if campaign_supp_selected else None
                    )
                    if validated_file_path:
                        with open(validated_file_path, "rb") as f:
                            st.success("✅ Validation complete!")
                            st.download_button("📥 Download Validated File", f, file_name="validated_output.xlsx")
                        os.remove(validated_file_path)
                        st.info("Validation columns updated in-place.")
                    else:
                        st.info("No file was generated due to validation errors. Please resolve the above issues and try again.")
                except Exception as e:
                    st.error(f"❌ Error while processing: {e}")
                finally:
                    # Clean up any temp suppression file we created from uploads
                    if suppression_temp_path and os.path.exists(suppression_temp_path):
                        try:
                            os.remove(suppression_temp_path)
                        except Exception:
                            pass

# --- FOOTER ---
st.markdown("""
<div style='text-align: center; font-size: 14px; color: #6c757d; margin-top:2.3rem; margin-bottom:0.4rem;'>
    © 2025 Interlink. All rights reserved.<br>
    Engineered for enterprise-grade data integrity and innovation by Felix Markas Salve.
</div>
""", unsafe_allow_html=True)