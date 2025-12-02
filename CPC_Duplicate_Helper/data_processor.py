import openpyxl
import streamlit as st
from datetime import datetime
from .duplicate_checker import DuplicateChecker
from .file_handler import FileHandler
from .utils import ensure_col_in_ws
from utils.file_utils import disqualify_lead

# Import internal checkers
from .internal_checkers import InternalCPCChecker, InternalDuplicateChecker, InternalPhoneChecker

class DataProcessor:
    """Main data processing coordinator - supports both internal and external validation"""
    
    def __init__(self):
        self.cpc_checker = None
        self.duplicate_checker = None
        self.stats = {
            'total_leads': 0,
            'cpc_violations': 0,
            'duplicates_found': 0,
            'internal_duplicates': 0,
            'internal_cpc_violations': 0,
            'internal_phone_conflicts': 0,
            'phone_conflicts': 0,
            'passed': 0,
            'permutation_errors': 0,
            'companies_checked': set(),
            'internal_companies_checked': set(),
            'processing_time': 0
        }
    
    def process_files_internal(self, lead_file, mapping, checks, cpc_limit=3):
        """Process single file for internal validation (first delivery mode)"""
        start_time = datetime.now()
        
        # Load lead workbook only
        lead_wb = openpyxl.load_workbook(lead_file)
        lead_ws = lead_wb.active
        
        # Convert sheet to data
        lead_headers, lead_data = FileHandler.sheet_to_dict_list(lead_ws)
        
        self.stats['total_leads'] = len(lead_data)
        
        # Setup progress tracking
        total_to_process = len(lead_data)
        show_progress = total_to_process > 1000
        if show_progress:
            progress_bar = st.progress(0)
            progress_text = st.empty()
        
        # Ensure required columns exist
        lead_status_col = ensure_col_in_ws(lead_headers, lead_ws, "Lead Status")
        dq_reason_col = ensure_col_in_ws(lead_headers, lead_ws, "DQ Reason")
        qa_comment_col = ensure_col_in_ws(lead_headers, lead_ws, "QA Comment")
        
        # Phone conflict column (always create if phone check is enabled)
        phone_conflict_col = None
        if checks['check_phone']:
            phone_conflict_col = ensure_col_in_ws(lead_headers, lead_ws, "Internal Phone Conflicts")
        
        # Run Internal CPC Check
        if checks['check_cpc']:
            st.info("🔍 Running Internal CPC Check...")
            internal_cpc_checker = InternalCPCChecker(cpc_limit)
            cpc_stats = internal_cpc_checker.run_internal_cpc_check(
                lead_data, lead_ws, mapping, lead_headers
            )
            self.stats.update(cpc_stats)
        
        # Run Internal Duplicate Check
        if checks['check_duplicates']:
            st.info("🔍 Running Internal Duplicate Check...")
            internal_duplicate_checker = InternalDuplicateChecker()
            dup_stats = internal_duplicate_checker.run_internal_duplicate_check(
                lead_data, lead_ws, mapping,
                lead_status_col, dq_reason_col, qa_comment_col
            )
            self.stats.update(dup_stats)
        
        # Run Internal Phone Conflict Check
        if checks['check_phone']:
            st.info("📞 Running Internal Phone Conflict Check...")
            internal_phone_checker = InternalPhoneChecker()
            phone_stats = internal_phone_checker.run_internal_phone_check(
                lead_data, lead_ws, mapping, phone_conflict_col
            )
            self.stats['internal_phone_conflicts'] = phone_stats['internal_phone_conflicts']
            self.stats['internal_phone_details'] = phone_stats['internal_phone_conflict_details']
        
        # Count passed leads
        for i in range(2, lead_ws.max_row + 1):
            if lead_ws.cell(i, lead_status_col).value != "Disqualified":
                self.stats['passed'] += 1
        
        # Clear progress indicators
        if show_progress:
            progress_bar.empty()
            progress_text.empty()
        
        # Calculate processing time
        self.stats['processing_time'] = (datetime.now() - start_time).total_seconds()
        
        return lead_wb, self.stats
    
    def process_files(self, delivery_file, lead_file, mapping, checks, cpc_limit=3):
        """Main processing function for external validation (subsequent delivery mode)"""
        start_time = datetime.now()
        
        # Load workbooks
        delivery_wb = openpyxl.load_workbook(delivery_file)
        delivery_ws = delivery_wb.active
        lead_wb = openpyxl.load_workbook(lead_file)
        lead_ws = lead_wb.active

        # Convert sheets to data
        delivery_headers, delivery_data = FileHandler.sheet_to_dict_list(delivery_ws)
        lead_headers, lead_data = FileHandler.sheet_to_dict_list(lead_ws)
        
        self.stats['total_leads'] = len(lead_data)
        
        # Setup progress tracking
        total_to_process = len(lead_data)
        show_progress = total_to_process > 1000
        if show_progress:
            progress_bar = st.progress(0)
            progress_text = st.empty()

        # Ensure required columns exist
        lead_status_col = ensure_col_in_ws(lead_headers, lead_ws, "Lead Status")
        dq_reason_col = ensure_col_in_ws(lead_headers, lead_ws, "DQ Reason")
        qa_comment_col = ensure_col_in_ws(lead_headers, lead_ws, "QA Comment")
        
        # Phone conflict column (always create if phone check is enabled)
        phone_conflict_col = None
        if checks['check_phone']:
            phone_conflict_col = ensure_col_in_ws(lead_headers, lead_ws, "Phone Conflicts")

        # Initialize checkers
        if checks['check_cpc']:
            from .domain_based_checker import DomainBasedChecker
            self.cpc_checker = DomainBasedChecker(cpc_limit)
            
        if checks['check_duplicates']:
            from .duplicate_checker import DuplicateChecker
            self.duplicate_checker = DuplicateChecker()

        # Run External Domain-Based CPC Check
        if checks['check_cpc']:
            st.info("🔍 Running External CPC Check...")
            cpc_stats = self.cpc_checker.run_domain_based_cpc_check(
                delivery_data, lead_data, lead_ws, mapping, lead_headers
            )
            self.stats.update(cpc_stats)

        # Run External Duplicate Check
        if checks['check_duplicates']:
            st.info("🔍 Running External Duplicate Check...")
            dup_stats = self.duplicate_checker.run_duplicate_check(
                delivery_data, lead_data, lead_ws, mapping,
                lead_status_col, dq_reason_col, qa_comment_col
            )
            self.stats.update(dup_stats)

        # Run External Phone Conflict Check (simplified - just check delivery vs lead)
        if checks['check_phone']:
            st.info("📞 Running External Phone Conflict Check...")
            
            from .simple_phone_checker import SimplePhoneChecker
            simple_phone_checker = SimplePhoneChecker()
            
            # Build phone mapping from delivery file
            simple_phone_checker.build_delivery_phone_map(
                delivery_data,
                mapping.get('delivery_phone', 'Not Available'),
                mapping.get('delivery_company', 'Not Available'),
                mapping.get('delivery_domain', 'Not Available')
            )
            
            # Check phone conflicts in lead file
            simple_phone_checker.check_phone_conflicts(
                lead_data, lead_ws,
                mapping.get('lead_phone', 'Not Available'),
                mapping.get('lead_company', 'Not Available'),
                mapping.get('lead_domain', 'Not Available'),
                phone_conflict_col
            )
            
            phone_stats = simple_phone_checker.get_stats()
            self.stats['phone_conflicts'] = phone_stats['phone_conflicts']
            self.stats['simple_phone_details'] = phone_stats['phone_conflict_details']

        # Run Internal Validation as well (for subsequent delivery mode)
        st.info("🔍 Running Additional Internal Validation...")
        
        # Internal CPC Check
        if checks['check_cpc']:
            internal_cpc_checker = InternalCPCChecker(cpc_limit)
            internal_cpc_stats = internal_cpc_checker.run_internal_cpc_check(
                lead_data, lead_ws, mapping, lead_headers
            )
            self.stats['internal_cpc_violations'] = internal_cpc_stats['internal_cpc_violations']
            self.stats['internal_companies_checked'] = internal_cpc_stats['internal_companies_checked']
        
        # Internal Duplicate Check
        if checks['check_duplicates']:
            internal_duplicate_checker = InternalDuplicateChecker()
            
            # Create a copy of mapping for internal checks (remove delivery prefixes)
            internal_mapping = {k.replace('delivery_', 'lead_'): v for k, v in mapping.items() if k.startswith('lead_')}
            
            internal_dup_stats = internal_duplicate_checker.run_internal_duplicate_check(
                lead_data, lead_ws, internal_mapping,
                lead_status_col, dq_reason_col, qa_comment_col
            )
            self.stats['internal_duplicates'] += internal_dup_stats['internal_duplicates']
            if 'internal_duplicate_details' not in self.stats:
                self.stats['internal_duplicate_details'] = []
            self.stats['internal_duplicate_details'].extend(internal_dup_stats['internal_duplicate_details'])
        
        # Internal Phone Conflict Check
        if checks['check_phone']:
            internal_phone_col = ensure_col_in_ws(lead_headers, lead_ws, "Internal Phone Conflicts")
            internal_phone_checker = InternalPhoneChecker()
            
            internal_phone_stats = internal_phone_checker.run_internal_phone_check(
                lead_data, lead_ws, mapping, internal_phone_col
            )
            self.stats['internal_phone_conflicts'] = internal_phone_stats['internal_phone_conflicts']
            self.stats['internal_phone_details'] = internal_phone_stats['internal_phone_conflict_details']

        # Count passed leads
        for i in range(2, lead_ws.max_row + 1):
            if lead_ws.cell(i, lead_status_col).value != "Disqualified":
                self.stats['passed'] += 1

        # Clear progress indicators
        if show_progress:
            progress_bar.empty()
            progress_text.empty()

        # Calculate processing time
        self.stats['processing_time'] = (datetime.now() - start_time).total_seconds()

        # Close original workbooks
        delivery_wb.close()
        
        return lead_wb, self.stats

    def get_comprehensive_stats(self):
        """Get all statistics including internal and external validation results"""
        stats = self.stats.copy()
        
        # Add domain analysis if CPC checker is available
        if self.cpc_checker and hasattr(self.cpc_checker, 'get_domain_analysis'):
            domain_analysis = self.cpc_checker.get_domain_analysis()
            stats['domain_analysis'] = domain_analysis
            
        return stats