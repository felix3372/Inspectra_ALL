import openpyxl
import streamlit as st
import pandas as pd

class FileHandler:
    """Handle file operations and data extraction"""
    
    @staticmethod
    @st.cache_data
    def get_headers_from_upload(file):
        """Get headers from uploaded file with caching"""
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else f"Column {i+1}" 
                      for i, cell in enumerate(ws[1])]
            
            # Get file stats
            row_count = ws.max_row - 1  # Exclude header
            col_count = ws.max_column
            
            wb.close()
            return headers, row_count, col_count
        except Exception as e:
            st.error(f"Error reading file: {str(e)}")
            return [], 0, 0

    @staticmethod
    def get_preview_data(file, num_rows=5):
        """Get preview of first few rows"""
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            headers = [str(cell.value).strip() if cell.value else f"Column {i+1}" 
                      for i, cell in enumerate(ws[1])]
            
            preview_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(num_rows+1, ws.max_row), values_only=True), start=2):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = str(value)[:50] if value else ""  # Truncate long values
                preview_data.append(row_dict)
            
            wb.close()
            return pd.DataFrame(preview_data)
        except Exception:
            return pd.DataFrame()

    @staticmethod
    def sheet_to_dict_list(ws):
        """Convert worksheet to list of dictionaries for faster processing"""
        headers = [str(cell.value).strip() if cell.value else f"Column {i+1}" 
                  for i, cell in enumerate(ws[1])]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row:  # Check if row exists
                row_list = list(row) if row else []
                row_list.extend([""] * max(0, len(headers) - len(row_list)))
                row_dict = {headers[i]: (str(row_list[i]).strip() if row_list[i] is not None else "") 
                           for i in range(len(headers))}
                data.append(row_dict)
        return headers, data